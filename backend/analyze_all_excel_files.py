#!/usr/bin/env python3
"""
Comprehensive Excel File Analysis for MARGEN.AI Integration
Analyzes all Excel files including sheets, columns, relationships, formulas, and correlations
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
import json
from collections import defaultdict

EXCEL_FOLDER = "/Users/inder/projects/mantrix-unified-nexxt-v1/excelfolder"

def analyze_excel_file(filepath):
    """Comprehensive analysis of an Excel file"""
    print(f"\n{'='*100}")
    print(f"ANALYZING: {os.path.basename(filepath)}")
    print(f"{'='*100}")

    analysis = {
        'filename': os.path.basename(filepath),
        'filepath': filepath,
        'sheets': {},
        'formulas': [],
        'potential_keys': [],
        'relationships': []
    }

    try:
        # Load workbook with openpyxl to detect formulas
        wb = openpyxl.load_workbook(filepath, data_only=False)

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n📊 SHEET: {sheet_name}")
            print("-" * 100)

            ws = wb[sheet_name]

            # Read with pandas for data analysis
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name)

                sheet_info = {
                    'row_count': len(df),
                    'columns': list(df.columns),
                    'column_count': len(df.columns),
                    'formulas': [],
                    'sample_data': {},
                    'dtypes': {},
                    'potential_keys': [],
                    'vlookup_references': []
                }

                # Basic info
                print(f"   Rows: {len(df):,}")
                print(f"   Columns: {len(df.columns)}")
                print(f"\n   📋 COLUMNS:")
                for i, col in enumerate(df.columns, 1):
                    dtype = str(df[col].dtype)
                    non_null = df[col].notna().sum()
                    unique = df[col].nunique()

                    sheet_info['dtypes'][col] = dtype

                    # Check if potential key (high uniqueness)
                    if unique == len(df) and unique > 1:
                        sheet_info['potential_keys'].append(col)
                        key_marker = " 🔑 [POTENTIAL KEY]"
                    elif unique / len(df) > 0.8:
                        key_marker = " 🔑 [HIGH UNIQUENESS]"
                    else:
                        key_marker = ""

                    print(f"      {i}. {col} ({dtype}) - {non_null:,}/{len(df):,} non-null, {unique:,} unique{key_marker}")

                    # Sample values
                    sample_vals = df[col].dropna().head(3).tolist()
                    sheet_info['sample_data'][col] = sample_vals
                    if sample_vals:
                        print(f"         Sample: {sample_vals}")

                # Detect formulas in openpyxl
                print(f"\n   🔍 FORMULAS DETECTED:")
                formula_count = 0
                for row in ws.iter_rows(min_row=2, max_row=min(100, ws.max_row)):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula = cell.value
                            formula_count += 1

                            # Detect VLOOKUP
                            if 'VLOOKUP' in formula.upper() or 'XLOOKUP' in formula.upper():
                                sheet_info['vlookup_references'].append({
                                    'cell': cell.coordinate,
                                    'formula': formula
                                })
                                print(f"      VLOOKUP at {cell.coordinate}: {formula[:100]}")

                            # Detect references to other sheets
                            if '!' in formula:
                                sheet_info['formulas'].append({
                                    'cell': cell.coordinate,
                                    'formula': formula[:200]
                                })
                                print(f"      Cross-sheet at {cell.coordinate}: {formula[:100]}")

                if formula_count == 0:
                    print(f"      No formulas detected (data-only sheet)")
                else:
                    print(f"      Total formulas: {formula_count}")

                analysis['sheets'][sheet_name] = sheet_info

            except Exception as e:
                print(f"   ❌ Error reading sheet: {str(e)}")
                analysis['sheets'][sheet_name] = {'error': str(e)}

        wb.close()

    except Exception as e:
        print(f"❌ Error analyzing file: {str(e)}")
        analysis['error'] = str(e)

    return analysis

def find_relationships(all_analyses):
    """Identify relationships between files based on common columns"""
    print(f"\n{'='*100}")
    print("🔗 IDENTIFYING RELATIONSHIPS BETWEEN FILES")
    print(f"{'='*100}")

    # Collect all columns from all sheets
    column_index = defaultdict(list)  # column_name -> [(filename, sheet_name)]

    for analysis in all_analyses:
        filename = analysis['filename']
        for sheet_name, sheet_info in analysis.get('sheets', {}).items():
            if 'columns' in sheet_info:
                for col in sheet_info['columns']:
                    column_index[col].append((filename, sheet_name))

    # Find common columns (potential join keys)
    print("\n📌 COMMON COLUMNS (Potential Join Keys):")
    relationships = []

    for col, occurrences in sorted(column_index.items()):
        if len(occurrences) > 1:
            print(f"\n   🔗 '{col}' appears in {len(occurrences)} locations:")
            for filename, sheet in occurrences:
                print(f"      - {filename} → {sheet}")

            relationships.append({
                'column': col,
                'occurrences': occurrences
            })

    return relationships

def suggest_integration_strategy(all_analyses, relationships):
    """Suggest how to integrate all files into MARGEN.AI"""
    print(f"\n{'='*100}")
    print("💡 INTEGRATION STRATEGY FOR MARGEN.AI")
    print(f"{'='*100}")

    print("\n🎯 RECOMMENDED INTEGRATION APPROACH:")

    # Categorize files
    print("\n1️⃣ CORE TRANSACTIONAL DATA (Already Loaded):")
    print("   ✅ csg.xlsx → fact_transactions table")

    print("\n2️⃣ MASTER/REFERENCE DATA (Enrich existing data):")
    master_files = [
        "#2 - Manufacturing Std Cost.xlsx",
        "#3 - Item Data File.xlsx",
        "2025 Territories - 6 Regions.xlsx"
    ]
    for f in master_files:
        print(f"   📊 {f}")
        analysis = next((a for a in all_analyses if a['filename'] == f), None)
        if analysis:
            for sheet_name, sheet_info in analysis.get('sheets', {}).items():
                if 'potential_keys' in sheet_info and sheet_info['potential_keys']:
                    print(f"      → Sheet: {sheet_name}")
                    print(f"         Keys: {', '.join(sheet_info['potential_keys'])}")

    print("\n3️⃣ INVOICE/BILLING DATA (Validate against transactions):")
    print("   📊 #1 - Invoice Data.xlsx")

    print("\n4️⃣ REGIONAL/TERRITORY DATA (Geographic analytics):")
    print("   📊 #4 - 6 Region - 2025 MSR - Tab 2025 Data.xlsx")
    print("   📊 2025 Territories - 6 Regions.xlsx")

    print("\n5️⃣ PERFORMANCE/COMMISSION DATA (Distributor analytics):")
    print("   📊 Cibolo Spine (Turgon) 2025 Commission (2).xlsx")
    print("   📊 Leap LLC (Knickerbocker) 2025 Commission (2).xlsx")
    print("   📊 SOP 6.0-01-10 Distributor Profitability Q3 2025 (3).xlsx")

    print("\n6️⃣ ANALYSIS/REVIEW DATA (Validation & insights):")
    print("   📊 CGS Review - ASP - System, Units, Facility - '25 8-20-25 9-5.xlsx")

    print("\n\n📋 PROPOSED NEW DATABASE TABLES:")
    print("\n   dim_items (Master item data)")
    print("      - item_code (PK)")
    print("      - item_description")
    print("      - category")
    print("      - unit_cost")
    print("      - unit_price")

    print("\n   dim_territories (Geographic hierarchy)")
    print("      - territory_id (PK)")
    print("      - territory_name")
    print("      - region")
    print("      - rep_name")

    print("\n   fact_invoices (Invoice transactions)")
    print("      - invoice_number (PK)")
    print("      - invoice_date")
    print("      - customer")
    print("      - amount")
    print("      - Join to fact_transactions on date/customer")

    print("\n   fact_commissions (Distributor commissions)")
    print("      - distributor")
    print("      - period")
    print("      - commission_amount")
    print("      - revenue")

    print("\n\n🎨 NEW MARGEN.AI COMPONENTS/ENHANCEMENTS:")
    print("\n   ✨ Territory Performance Dashboard")
    print("      - Revenue by territory")
    print("      - Rep performance rankings")
    print("      - Geographic heatmaps")

    print("\n   ✨ Item/Product Analytics")
    print("      - Item-level profitability")
    print("      - Product mix analysis")
    print("      - Pricing optimization")

    print("\n   ✨ Commission & Profitability")
    print("      - Distributor commission tracking")
    print("      - Net profitability after commissions")
    print("      - Commission vs. margin analysis")

    print("\n   ✨ Invoice Reconciliation")
    print("      - Invoice vs. transaction matching")
    print("      - Outstanding receivables")
    print("      - Payment tracking")

def main():
    """Main analysis function"""
    print("="*100)
    print("COMPREHENSIVE EXCEL FILE ANALYSIS FOR MARGEN.AI")
    print("="*100)

    excel_files = [
        "csg.xlsx",
        "#1 - Invoice Data.xlsx",
        "#2 - Manufacturing Std Cost.xlsx",
        "#3 - Item Data File.xlsx",
        "#4 - 6 Region - 2025 MSR - Tab 2025 Data.xlsx",
        "2025 Territories - 6 Regions.xlsx",
        "CGS Review - ASP - System, Units, Facility - '25 8-20-25 9-5.xlsx",
        "Cibolo Spine (Turgon) 2025 Commission (2).xlsx",
        "Leap LLC (Knickerbocker) 2025 Commission (2).xlsx",
        "SOP 6.0-01-10 Distributor Profitability Q3 2025 (3).xlsx"
    ]

    all_analyses = []

    for filename in excel_files:
        filepath = os.path.join(EXCEL_FOLDER, filename)
        if os.path.exists(filepath):
            analysis = analyze_excel_file(filepath)
            all_analyses.append(analysis)
        else:
            print(f"\n⚠️  File not found: {filename}")

    # Find relationships
    relationships = find_relationships(all_analyses)

    # Suggest integration strategy
    suggest_integration_strategy(all_analyses, relationships)

    # Save detailed analysis to JSON
    output_file = "/Users/inder/projects/mantrix-unified-nexxt-v1/backend/excel_analysis_detailed.json"
    with open(output_file, 'w') as f:
        json.dump({
            'analyses': all_analyses,
            'relationships': relationships
        }, f, indent=2, default=str)

    print(f"\n\n💾 Detailed analysis saved to: {output_file}")
    print("\n" + "="*100)
    print("ANALYSIS COMPLETE")
    print("="*100 + "\n")

if __name__ == "__main__":
    main()
