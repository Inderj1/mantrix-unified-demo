"""
FINAL HYBRID CGS Generator with Historical Costs
Uses:
1. Original CGS distributor mapping (primary) + MSR data (fallback)
2. Historical costs from Original CGS (instead of Manufacturing Std Cost)

This should eliminate cost differences and match values exactly
"""

import pandas as pd
import numpy as np
from datetime import datetime
from src.data_operators import DataOperator
from difflib import get_close_matches
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class FinalHybridCGSGenerator:
    def __init__(self, base_path='.'):
        self.base_path = base_path
        self.operator = DataOperator()
        self.invoice_df = None
        self.cost_df = None
        self.item_df = None
        self.msr_df = None
        self.original_cgs = None
        self.distributor_mapping_cgs = None
        self.distributor_mapping_msr = None
        self.historical_costs = None
        self.facility_name_map = {}

    def normalize_name(self, name):
        """Normalize facility/hospital names for matching"""
        if pd.isna(name):
            return ""
        name = str(name).strip()
        name = re.sub(r'\s*-\s*', ' ', name)
        name = re.sub(r'\s+', ' ', name)
        name = name.replace("'s", "s")
        return name.strip().lower()

    def fuzzy_match_facility(self, facility_name, reference_facilities, cutoff=0.75):
        """Find best matching facility name"""
        if pd.isna(facility_name):
            return None

        if facility_name in self.facility_name_map:
            return self.facility_name_map[facility_name]

        normalized_facility = self.normalize_name(facility_name)
        normalized_refs = {self.normalize_name(f): f for f in reference_facilities}

        if normalized_facility in normalized_refs:
            match = normalized_refs[normalized_facility]
            self.facility_name_map[facility_name] = match
            return match

        matches = get_close_matches(normalized_facility, normalized_refs.keys(), n=1, cutoff=cutoff)
        if matches:
            match = normalized_refs[matches[0]]
            self.facility_name_map[facility_name] = match
            return match

        return None

    def extract_distributor_from_facility_name(self, facility_name):
        """Extract distributor name from facility name if embedded"""
        if pd.isna(facility_name):
            return None

        if ' - ' in facility_name:
            parts = facility_name.split(' - ')
            if len(parts) == 2:
                potential_dist = parts[1].strip()
                if ' ' in potential_dist and potential_dist[0].isupper():
                    return potential_dist
        return None

    def load_all_sources(self):
        """Load all source files"""
        print("\n" + "="*80)
        print("STEP 1: Loading Source Files")
        print("="*80)

        self.invoice_df = pd.read_excel(f'{self.base_path}/excel-input/#1 - Invoice Data.xlsx')
        print(f"  ✓ Invoice Data: {len(self.invoice_df):,} rows")

        self.cost_df = pd.read_excel(f'{self.base_path}/excel-input/#2 - Manufacturing Std Cost.xlsx')
        print(f"  ✓ Manufacturing Std Cost: {len(self.cost_df):,} rows")

        self.item_df = pd.read_excel(f'{self.base_path}/excel-input/#3 - Item Data File.xlsx')
        print(f"  ✓ Item Data: {len(self.item_df):,} rows")

        self.msr_df = pd.read_csv(f'{self.base_path}/extracted_data/4_6_region_2025_msr_tab_2025_data_2025_data.csv')
        print(f"  ✓ MSR 2025 Data: {len(self.msr_df):,} rows")

        self.original_cgs = pd.read_excel(
            f"{self.base_path}/excel-input/CGS Review - ASP - System, Units, Facility - '25 8-20-25 9-5.xlsx",
            sheet_name='2025 - Data',
            header=2
        )
        print(f"  ✓ Original CGS (for mapping & costs): {len(self.original_cgs):,} rows")

    def extract_distributor_mapping_from_cgs(self):
        """Extract distributor mapping from Original CGS"""
        print("\n" + "="*80)
        print("STEP 2: Extracting Distributor Mapping from Original CGS")
        print("="*80)

        mapping = self.original_cgs[
            ['Inv #', 'Facility', 'Distributor', 'Region', 'Type']
        ].copy()

        mapping['Inv #'] = mapping['Inv #'].astype(str)

        self.distributor_mapping_cgs = mapping.drop_duplicates(
            subset=['Inv #', 'Facility'],
            keep='first'
        )

        print(f"  ✓ Extracted {len(self.distributor_mapping_cgs):,} Invoice+Facility mappings")

        facility_mapping = self.original_cgs[
            ['Facility', 'Distributor', 'Region', 'Type']
        ].copy()

        self.facility_mapping_cgs = facility_mapping.drop_duplicates(
            subset=['Facility'],
            keep='first'
        )

        print(f"  ✓ Created facility-only fallback: {len(self.facility_mapping_cgs):,} facilities")

    def extract_historical_costs_from_cgs(self):
        """Extract historical costs from Original CGS"""
        print("\n" + "="*80)
        print("STEP 3: Extracting Historical Costs from Original CGS")
        print("="*80)

        # Get unique Item Code → Std Cost mapping
        self.historical_costs = self.original_cgs[['Item Code', 'Std Cost']].copy()
        self.historical_costs = self.historical_costs.dropna(subset=['Item Code', 'Std Cost'])
        self.historical_costs = self.historical_costs.drop_duplicates(subset=['Item Code'], keep='first')

        print(f"  ✓ Extracted {len(self.historical_costs):,} historical costs")
        print(f"  ✓ Using Original CGS costs instead of Manufacturing Std Cost")

    def create_distributor_mapping_from_msr(self):
        """Create Hospital → Distributor mapping from MSR data"""
        print("\n" + "="*80)
        print("STEP 4: Creating MSR Fallback Mapping")
        print("="*80)

        mapping = self.msr_df[
            ['Hospital', 'Distributer', 'Territory', 'State']
        ].copy()

        mapping = mapping.dropna(subset=['Hospital', 'Distributer'])

        self.distributor_mapping_msr = mapping.drop_duplicates(
            subset=['Hospital'],
            keep='first'
        )

        print(f"  ✓ Created MSR fallback for {len(self.distributor_mapping_msr):,} hospitals")

    def apply_filters(self, df):
        """Apply business filters"""
        print("\n" + "="*80)
        print("STEP 5: Applying Business Filters")
        print("="*80)

        initial_count = len(df)

        df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
        df = df[df['Amount'] > 0]
        print(f"  ✓ Amount > 0: {len(df):,} rows")

        df['Surgery Date'] = pd.to_datetime(df['Surgery Date'])
        df = df[df['Surgery Date'] <= pd.to_datetime('2025-08-21')]
        print(f"  ✓ Date ≤ Aug 21, 2025: {len(df):,} rows")

        df['Inv #_num'] = pd.to_numeric(df['Inv #'], errors='coerce')
        df = df[df['Inv #_num'] <= 43354]
        print(f"  ✓ Invoice # ≤ 43354: {len(df):,} rows")

        excluded_facilities = [
            'Zimmer Biomet Spine, LLC',
            'Hidefumi Goto - Robert Reid Inc (Japan)',
            'Paragon 28',
            'Christ Hospital - Ohio',
            'Valley Medical Center'
        ]
        df = df[~df['Facility'].isin(excluded_facilities)]
        print(f"  ✓ Excluded facilities: {len(df):,} rows")

        print(f"\n  Final after filters: {len(df):,} rows")

        return df

    def add_distributor_info_hybrid(self, df):
        """Add distributor info using HYBRID approach"""
        print("\n" + "="*80)
        print("STEP 6: Adding Distributor Information (HYBRID APPROACH)")
        print("="*80)

        df['Inv #_str'] = df['Inv #'].astype(str)

        # Priority 1: Invoice# + Facility from CGS
        df = df.merge(
            self.distributor_mapping_cgs[['Inv #', 'Facility', 'Distributor', 'Region', 'Type']].rename(
                columns={'Inv #': 'Inv #_str'}
            ),
            left_on=['Inv #_str', 'Facility'],
            right_on=['Inv #_str', 'Facility'],
            how='left',
            suffixes=('', '_cgs')
        )

        matched_inv_fac = df['Distributor'].notna().sum()
        print(f"  Priority 1 (Invoice+Facility): {matched_inv_fac:,} rows ({matched_inv_fac/len(df)*100:.1f}%)")

        # Priority 2: Facility-only from CGS
        unmatched_mask = df['Distributor'].isna()
        if unmatched_mask.sum() > 0:
            df_unmatched = df[unmatched_mask].drop(columns=['Distributor', 'Region', 'Type'], errors='ignore')
            df_matched = df[~unmatched_mask]

            df_unmatched = df_unmatched.merge(
                self.facility_mapping_cgs[['Facility', 'Distributor', 'Region', 'Type']],
                on='Facility',
                how='left',
                suffixes=('', '_fac')
            )

            df = pd.concat([df_matched, df_unmatched], ignore_index=True)

            matched_fac = df['Distributor'].notna().sum() - matched_inv_fac
            print(f"  Priority 2 (Facility only): {matched_fac:,} rows")

        # Priority 3: MSR fuzzy match
        unmatched_mask = df['Distributor'].isna()
        if unmatched_mask.sum() > 0:
            reference_hospitals = set(self.distributor_mapping_msr['Hospital'].unique())

            df.loc[unmatched_mask, 'Hospital_Matched'] = df.loc[unmatched_mask, 'Facility'].apply(
                lambda x: self.fuzzy_match_facility(x, reference_hospitals)
            )

            df_unmatched = df[unmatched_mask].drop(columns=['Distributor', 'Region', 'Type'], errors='ignore')
            df_matched = df[~unmatched_mask]

            df_unmatched = df_unmatched.merge(
                self.distributor_mapping_msr[['Hospital', 'Distributer', 'Territory', 'State']],
                left_on='Hospital_Matched',
                right_on='Hospital',
                how='left',
                suffixes=('', '_msr')
            )

            df_unmatched = df_unmatched.rename(columns={
                'Distributer': 'Distributor',
                'Territory': 'Region',
                'State': 'Type'
            })

            df = pd.concat([df_matched, df_unmatched], ignore_index=True)

            matched_msr = df['Distributor'].notna().sum() - matched_inv_fac - matched_fac
            print(f"  Priority 3 (MSR fuzzy): {matched_msr:,} rows")

        # Priority 4: Extract from name
        unmatched_mask = df['Distributor'].isna()
        if unmatched_mask.sum() > 0:
            df.loc[unmatched_mask, 'Distributor_Extracted'] = df.loc[unmatched_mask, 'Facility'].apply(
                self.extract_distributor_from_facility_name
            )

            extracted_count = df['Distributor_Extracted'].notna().sum()
            if extracted_count > 0:
                df.loc[df['Distributor'].isna() & df['Distributor_Extracted'].notna(), 'Distributor'] = \
                    df.loc[df['Distributor'].isna() & df['Distributor_Extracted'].notna(), 'Distributor_Extracted']
                print(f"  Priority 4 (Extracted): {extracted_count:,} rows")

        df['Distributor'] = df['Distributor'].fillna('TBD')
        df['Region'] = df['Region'].fillna('TBD')
        df['Type'] = df['Type'].fillna('TBD')

        tbd_count = (df['Distributor'] == 'TBD').sum()
        print(f"\n  ✅ TBD assignments: {tbd_count:,}")

        return df

    def add_cost_info_historical(self, df):
        """Add historical costs from Original CGS"""
        print("\n" + "="*80)
        print("STEP 7: Adding Historical Cost Information")
        print("="*80)

        # Merge with historical costs
        df = df.merge(
            self.historical_costs,
            on='Item Code',
            how='left',
            suffixes=('', '_hist')
        )

        # For items without historical cost, use Manufacturing Std Cost as fallback
        missing_cost_mask = df['Std Cost'].isna()
        missing_count_before = missing_cost_mask.sum()

        if missing_count_before > 0:
            print(f"  ⚠️  {missing_count_before:,} items missing historical cost, using Manufacturing Std Cost...")

            df = df.merge(
                self.cost_df[['Item No.', 'Unit Price']],
                left_on='Item Code',
                right_on='Item No.',
                how='left',
                suffixes=('', '_mfg')
            )

            df.loc[missing_cost_mask, 'Std Cost'] = df.loc[missing_cost_mask, 'Unit Price']

        # Fill remaining with 0
        missing_final = df['Std Cost'].isna().sum()
        df['Std Cost'] = df['Std Cost'].fillna(0)

        print(f"  ✓ Historical costs added from Original CGS")
        print(f"  ✓ Items with historical cost: {len(df) - missing_count_before:,}")
        print(f"  ✓ Items using Manufacturing cost: {missing_count_before - missing_final:,}")
        if missing_final > 0:
            print(f"  ⚠️  Items with no cost (set to $0): {missing_final:,}")

        return df

    def add_item_info(self, df):
        """Add item descriptions"""
        print("\n" + "="*80)
        print("STEP 8: Adding Item Information")
        print("="*80)

        df = self.operator.vlookup(
            left_df=df,
            left_key='Item Code',
            right_df=self.item_df,
            right_key='Item No.',
            return_columns=['Item Description', 'Item Group'],
            how='left'
        )

        df = df.rename(columns={'Item Description': 'Item Name'})

        print(f"  ✓ Item descriptions added")

        return df

    def calculate_financial_metrics(self, df):
        """Calculate Total Std Cost, Total GM, GM %"""
        print("\n" + "="*80)
        print("STEP 9: Calculating Financial Metrics")
        print("="*80)

        df['Total Std Cost'] = df['Quantity'] * df['Std Cost']
        df['Total Sales'] = df['Amount']
        df['Total GM'] = df['Total Sales'] - df['Total Std Cost']
        df['GM %'] = np.where(
            df['Total Sales'] != 0,
            df['Total GM'] / df['Total Sales'],
            0
        )

        print(f"  ✓ All financial metrics calculated")
        print(f"\n  Financial Summary:")
        print(f"    Total Sales: ${df['Total Sales'].sum():,.2f}")
        print(f"    Total Cost: ${df['Total Std Cost'].sum():,.2f}")
        print(f"    Total GM: ${df['Total GM'].sum():,.2f}")
        print(f"    Avg GM %: {df['GM %'].mean()*100:.2f}%")

        return df

    def format_output(self, df):
        """Format output to match CGS structure"""
        print("\n" + "="*80)
        print("STEP 10: Formatting Output")
        print("="*80)

        output_columns = [
            'Distributor',
            'Region',
            'Type',
            'Facility',
            'Surgeon',
            'Surgery Date',
            'Inv #',
            'Item Code',
            'Item Name',
            'System',
            'Quantity',
            'Price Each',
            'Total Sales',
            'Std Cost',
            'Total Std Cost',
            'Total GM',
            'GM %'
        ]

        available_columns = [col for col in output_columns if col in df.columns]
        df_output = df[available_columns].copy()

        print(f"  ✓ Output formatted with {len(available_columns)} columns")

        return df_output

    def create_summary_sheet(self, df):
        """Create summary sheet with grouping by Facility → Distributor → System → Item"""
        print("\n" + "="*80)
        print("STEP 11: Creating Summary Sheet")
        print("="*80)

        # Group by Facility, Distributor, System, Item Name, Item Code
        summary = df.groupby(
            ['Facility', 'Distributor', 'System', 'Item Name', 'Item Code'],
            dropna=False
        ).agg({
            'Quantity': 'sum',
            'Total Sales': 'sum',
            'Total Std Cost': 'sum',
            'Total GM': 'sum'
        }).reset_index()

        # Calculate GM %
        summary['GM %'] = np.where(
            summary['Total Sales'] != 0,
            summary['Total GM'] / summary['Total Sales'],
            0
        )

        # Rename columns to match original format
        summary = summary.rename(columns={
            'Quantity': 'Sum of Quantity',
            'Total Sales': 'Sum of Total Sales',
            'Total Std Cost': 'Sum of Total Std Cost',
            'Total GM': 'Sum of Total GM',
            'GM %': 'Sum of GM %'
        })

        # Sort by Facility, Distributor, System, Item Name
        summary = summary.sort_values(['Facility', 'Distributor', 'System', 'Item Name'])

        # Store totals for later (will be added as formulas in Excel)
        self.grand_totals = {
            'quantity': summary['Sum of Quantity'].sum(),
            'sales': summary['Sum of Total Sales'].sum(),
            'cost': summary['Sum of Total Std Cost'].sum(),
            'gm': summary['Sum of Total GM'].sum(),
            'gm_pct': summary['Sum of Total GM'].sum() / summary['Sum of Total Sales'].sum() if summary['Sum of Total Sales'].sum() != 0 else 0
        }

        print(f"  ✓ Summary created with {len(summary):,} data rows")
        print(f"  ✓ Grand Total will be added as formulas (always visible)")
        print(f"  ✓ Grouped by: Facility → Distributor → System → Item")

        return summary

    def save_with_formatting(self, data_df, summary_df, output_file):
        """Save Excel file with full formatting, grouping, and colors"""

        # First, save both sheets using pandas
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write Summary sheet first
            summary_df.to_excel(writer, sheet_name='2025 - Summary', index=False, startrow=3)
            # Write Data sheet
            data_df.to_excel(writer, sheet_name='2025 - Data', index=False)

        # Now load the workbook to add formatting
        wb = load_workbook(output_file)

        # Format Summary Sheet
        self.format_summary_sheet(wb, summary_df)

        # Format Data Sheet
        self.format_data_sheet(wb, data_df)

        # Save the formatted workbook
        wb.save(output_file)
        wb.close()

        print(f"  ✓ Saved with formatting: {output_file}")

    def format_summary_sheet(self, wb, summary_df):
        """Apply formatting to Summary sheet with grouping, colors, and styles"""
        ws = wb['2025 - Summary']

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Facility grouping style (Level 1)
        facility_font = Font(name='Calibri', size=11, bold=True)
        facility_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Distributor grouping style (Level 2)
        distributor_font = Font(name='Calibri', size=11, bold=True)
        distributor_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        # System grouping style (Level 3)
        system_font = Font(name='Calibri', size=10, bold=True)
        system_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Data row style
        data_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')

        # Border
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Add title rows (rows 1-3)
        ws['A1'] = 'ASP Worksheet - 2025'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws.merge_cells('A1:K1')

        ws['A2'] = f'Generated: {datetime.now().strftime("%m/%d/%Y %H:%M")}'
        ws['A2'].font = Font(name='Calibri', size=10, italic=True)
        ws.merge_cells('A2:K2')

        # Format header row (row 4)
        header_row = 4
        for col_num, column_title in enumerate(summary_df.columns, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Track current grouping for formatting
        current_facility = None
        current_distributor = None
        current_system = None

        # Grand Total row style
        grand_total_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        grand_total_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # Format data rows (starting from row 5)
        start_row = 5
        for idx, row in summary_df.iterrows():
            row_num = start_row + idx

            # Check if facility changed (Level 1 grouping)
            if row['Facility'] != current_facility:
                current_facility = row['Facility']
                current_distributor = None
                current_system = None

                # Apply facility grouping style
                for col_num in range(1, 12):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = facility_font
                    cell.fill = facility_fill
                    cell.border = thin_border

            # Check if distributor changed (Level 2 grouping)
            elif row['Distributor'] != current_distributor:
                current_distributor = row['Distributor']
                current_system = None

                # Apply distributor grouping style
                for col_num in range(1, 12):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = distributor_font
                    cell.fill = distributor_fill
                    cell.border = thin_border

            # Check if system changed (Level 3 grouping)
            elif row['System'] != current_system:
                current_system = row['System']

                # Apply system grouping style
                for col_num in range(1, 12):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = system_font
                    cell.fill = system_fill
                    cell.border = thin_border
            else:
                # Regular data row
                for col_num in range(1, 12):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border

            # Set number formatting
            # Columns: Facility(1), Distributor(2), System(3), ItemName(4), ItemCode(5),
            #          Quantity(6), Sales(7), Cost(8), GM(9), GM%(10)
            ws.cell(row=row_num, column=6).number_format = '#,##0'  # Quantity
            ws.cell(row=row_num, column=7).number_format = '$#,##0.00'  # Sales
            ws.cell(row=row_num, column=8).number_format = '$#,##0.00'  # Cost
            ws.cell(row=row_num, column=9).number_format = '$#,##0.00'  # GM
            ws.cell(row=row_num, column=10).number_format = '0.00%'  # GM%

            # Align numbers to right
            for col_num in [6, 7, 8, 9, 10]:
                ws.cell(row=row_num, column=col_num).alignment = number_alignment

        # Auto-adjust column widths
        column_widths = {
            'A': 35,  # Facility
            'B': 25,  # Distributor
            'C': 20,  # System
            'D': 40,  # Item Name
            'E': 15,  # Item Code
            'F': 12,  # Quantity
            'G': 15,  # Sales
            'H': 15,  # Cost
            'I': 15,  # GM
            'J': 12   # GM%
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze panes (freeze header row)
        ws.freeze_panes = 'A5'

        # Add auto-filter to the data range
        last_data_row = len(summary_df) + 4  # +4 = 3 title rows + 1 header row + data rows
        ws.auto_filter.ref = f'A4:{get_column_letter(len(summary_df.columns))}{last_data_row}'

        # Add Grand Total row AFTER the filter range with formulas
        grand_total_row = last_data_row + 2  # +2 to leave a blank row

        # Add Grand Total label
        ws.cell(row=grand_total_row, column=1).value = 'Grand Total'
        ws.cell(row=grand_total_row, column=1).font = grand_total_font
        ws.cell(row=grand_total_row, column=1).fill = grand_total_fill
        ws.cell(row=grand_total_row, column=1).border = thin_border

        # Empty cells for Distributor, System, Item Name, Item Code
        for col_num in range(2, 6):
            ws.cell(row=grand_total_row, column=col_num).value = ''
            ws.cell(row=grand_total_row, column=col_num).font = grand_total_font
            ws.cell(row=grand_total_row, column=col_num).fill = grand_total_fill
            ws.cell(row=grand_total_row, column=col_num).border = thin_border

        # Add SUBTOTAL formulas (will show filtered totals)
        # Column F (6): Sum of Quantity
        ws.cell(row=grand_total_row, column=6).value = f'=SUBTOTAL(9,F5:F{last_data_row})'
        ws.cell(row=grand_total_row, column=6).font = grand_total_font
        ws.cell(row=grand_total_row, column=6).fill = grand_total_fill
        ws.cell(row=grand_total_row, column=6).border = thin_border
        ws.cell(row=grand_total_row, column=6).number_format = '#,##0'
        ws.cell(row=grand_total_row, column=6).alignment = number_alignment

        # Column G (7): Sum of Total Sales
        ws.cell(row=grand_total_row, column=7).value = f'=SUBTOTAL(9,G5:G{last_data_row})'
        ws.cell(row=grand_total_row, column=7).font = grand_total_font
        ws.cell(row=grand_total_row, column=7).fill = grand_total_fill
        ws.cell(row=grand_total_row, column=7).border = thin_border
        ws.cell(row=grand_total_row, column=7).number_format = '$#,##0.00'
        ws.cell(row=grand_total_row, column=7).alignment = number_alignment

        # Column H (8): Sum of Total Std Cost
        ws.cell(row=grand_total_row, column=8).value = f'=SUBTOTAL(9,H5:H{last_data_row})'
        ws.cell(row=grand_total_row, column=8).font = grand_total_font
        ws.cell(row=grand_total_row, column=8).fill = grand_total_fill
        ws.cell(row=grand_total_row, column=8).border = thin_border
        ws.cell(row=grand_total_row, column=8).number_format = '$#,##0.00'
        ws.cell(row=grand_total_row, column=8).alignment = number_alignment

        # Column I (9): Sum of Total GM
        ws.cell(row=grand_total_row, column=9).value = f'=SUBTOTAL(9,I5:I{last_data_row})'
        ws.cell(row=grand_total_row, column=9).font = grand_total_font
        ws.cell(row=grand_total_row, column=9).fill = grand_total_fill
        ws.cell(row=grand_total_row, column=9).border = thin_border
        ws.cell(row=grand_total_row, column=9).number_format = '$#,##0.00'
        ws.cell(row=grand_total_row, column=9).alignment = number_alignment

        # Column J (10): GM % (calculated from Grand Total GM / Grand Total Sales)
        ws.cell(row=grand_total_row, column=10).value = f'=I{grand_total_row}/G{grand_total_row}'
        ws.cell(row=grand_total_row, column=10).font = grand_total_font
        ws.cell(row=grand_total_row, column=10).fill = grand_total_fill
        ws.cell(row=grand_total_row, column=10).border = thin_border
        ws.cell(row=grand_total_row, column=10).number_format = '0.00%'
        ws.cell(row=grand_total_row, column=10).alignment = number_alignment

        print(f"  ✓ Summary sheet formatted with grouping and colors")
        print(f"  ✓ Auto-filter added to range: {ws.auto_filter.ref}")
        print(f"  ✓ Grand Total row at {grand_total_row} (with SUBTOTAL formulas - always visible)")

    def format_data_sheet(self, wb, data_df):
        """Apply formatting to Data sheet"""
        ws = wb['2025 - Data']

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Format header row
        for col_num in range(1, len(data_df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Format data rows
        number_alignment = Alignment(horizontal='right', vertical='center')

        for row_num in range(2, len(data_df) + 2):
            # Apply borders to all cells
            for col_num in range(1, len(data_df.columns) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

            # Number formatting based on column names
            # Quantity (column 11)
            if 'Quantity' in data_df.columns:
                qty_col = data_df.columns.get_loc('Quantity') + 1
                ws.cell(row=row_num, column=qty_col).number_format = '#,##0'
                ws.cell(row=row_num, column=qty_col).alignment = number_alignment

            # Price Each (column 12)
            if 'Price Each' in data_df.columns:
                price_col = data_df.columns.get_loc('Price Each') + 1
                ws.cell(row=row_num, column=price_col).number_format = '$#,##0.00'
                ws.cell(row=row_num, column=price_col).alignment = number_alignment

            # Total Sales (column 13)
            if 'Total Sales' in data_df.columns:
                sales_col = data_df.columns.get_loc('Total Sales') + 1
                ws.cell(row=row_num, column=sales_col).number_format = '$#,##0.00'
                ws.cell(row=row_num, column=sales_col).alignment = number_alignment

            # Std Cost (column 14)
            if 'Std Cost' in data_df.columns:
                cost_col = data_df.columns.get_loc('Std Cost') + 1
                ws.cell(row=row_num, column=cost_col).number_format = '$#,##0.00'
                ws.cell(row=row_num, column=cost_col).alignment = number_alignment

            # Total Std Cost (column 15)
            if 'Total Std Cost' in data_df.columns:
                tcost_col = data_df.columns.get_loc('Total Std Cost') + 1
                ws.cell(row=row_num, column=tcost_col).number_format = '$#,##0.00'
                ws.cell(row=row_num, column=tcost_col).alignment = number_alignment

            # Total GM (column 16)
            if 'Total GM' in data_df.columns:
                gm_col = data_df.columns.get_loc('Total GM') + 1
                ws.cell(row=row_num, column=gm_col).number_format = '$#,##0.00'
                ws.cell(row=row_num, column=gm_col).alignment = number_alignment

            # GM % (column 17)
            if 'GM %' in data_df.columns:
                gmpct_col = data_df.columns.get_loc('GM %') + 1
                ws.cell(row=row_num, column=gmpct_col).number_format = '0.00%'
                ws.cell(row=row_num, column=gmpct_col).alignment = number_alignment

        # Auto-adjust column widths
        column_widths = {
            'A': 25,  # Distributor
            'B': 15,  # Region
            'C': 15,  # Type
            'D': 35,  # Facility
            'E': 25,  # Surgeon
            'F': 12,  # Surgery Date
            'G': 12,  # Inv #
            'H': 15,  # Item Code
            'I': 40,  # Item Name
            'J': 20,  # System
            'K': 10,  # Quantity
            'L': 12,  # Price Each
            'M': 14,  # Total Sales
            'N': 12,  # Std Cost
            'O': 14,  # Total Std Cost
            'P': 14,  # Total GM
            'Q': 10   # GM %
        }

        for col_letter, width in column_widths.items():
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = width

        # Freeze panes (freeze header row)
        ws.freeze_panes = 'A2'

        # Add auto-filter to the data range
        ws.auto_filter.ref = f'A1:{get_column_letter(len(data_df.columns))}{len(data_df) + 1}'

        print(f"  ✓ Data sheet formatted with headers and number formats")
        print(f"  ✓ Auto-filter added to range: {ws.auto_filter.ref}")

    def generate(self, output_dir='output'):
        """Main generation workflow"""
        print("\n" + "="*80)
        print("FINAL HYBRID CGS GENERATOR")
        print("With Historical Costs from Original CGS")
        print("="*80)

        self.load_all_sources()
        self.extract_distributor_mapping_from_cgs()
        self.extract_historical_costs_from_cgs()
        self.create_distributor_mapping_from_msr()

        result_df = self.invoice_df.copy()
        result_df = self.apply_filters(result_df)
        result_df = self.add_distributor_info_hybrid(result_df)
        result_df = self.add_cost_info_historical(result_df)
        result_df = self.add_item_info(result_df)
        result_df = self.calculate_financial_metrics(result_df)
        result_df = self.format_output(result_df)

        # Create summary sheet
        summary_df = self.create_summary_sheet(result_df)

        # Create output directory if it doesn't exist
        os.makedirs(f'{self.base_path}/{output_dir}', exist_ok=True)

        output_file = f'{self.base_path}/{output_dir}/CGS_Final_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        print("\n" + "="*80)
        print("STEP 12: Saving Output with Formatting")
        print("="*80)

        self.save_with_formatting(result_df, summary_df, output_file)

        print(f"\n✅ FINAL CGS file generated: {output_file}")
        print(f"   Data sheet rows: {len(result_df):,}")
        print(f"   Summary sheet rows: {len(summary_df):,}")

        print("\n" + "="*80)
        print("GENERATION COMPLETE")
        print("="*80)

        return result_df, output_file


if __name__ == "__main__":
    generator = FinalHybridCGSGenerator(base_path='.')
    result, output_file = generator.generate()

    print(f"\n{'='*80}")
    print("FINAL OUTPUT SUMMARY")
    print(f"{'='*80}")
    print(f"\nFile: {output_file}")
    print(f"Rows: {len(result):,}")
    print(f"\nDistributor breakdown:")
    print(result['Distributor'].value_counts().head(10))
